from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.views import LoginView
from calendar import monthcalendar
from datetime import datetime, date, timedelta
import csv
import logging
from .models import Staff, Holiday, DutySchedule, DutyOrder, User, DutySwapRequest
from django.db.models import Max
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side,
    NamedStyle
)

logger = logging.getLogger(__name__)

def is_admin(user):
    return user.is_staff or user.is_superuser

class CustomLoginView(LoginView):
    template_name = 'duty/login.html'
    
@login_required
def duty_calendar(request):
    current_date = datetime.now()
    year = current_date.year
    month = current_date.month
    
    # 获取月历
    cal = monthcalendar(year, month)
    
    # 获取当月所有排班和节假日
    schedules = DutySchedule.objects.filter(
        date__year=year,
        date__month=month
    ).select_related('staff__user')
    
    holidays = Holiday.objects.filter(
        date__year=year,
        date__month=month
    ).values_list('date__day', flat=True)
    
    # 转换为字典格式
    schedule_dict = {
        schedule.date.day: (schedule.staff.user.username, schedule.is_holiday)
        for schedule in schedules
    }
    
    # 获取当前用户收到的待处理换班申请
    pending_swaps = []
    if request.user.is_authenticated:
        staff = Staff.objects.filter(user=request.user).first()
        if staff:
            pending_swaps = DutySwapRequest.objects.filter(
                target__user=request.user,
                status='pending'
            ).select_related(
                'requester__user',
                'requester_duty',
                'target_duty'
            )
    
    context = {
        'calendar': cal,
        'current_date': current_date,
        'schedule_dict': schedule_dict,
        'holidays': list(holidays),
        'is_admin': request.user.is_staff,
        'pending_swaps': pending_swaps,
        'today': date.today().strftime('%Y-%m-%d'),  # 格式化今天的日期为字符串
    }
    
    return render(request, 'duty/calendar.html', context)

@require_POST
@user_passes_test(is_admin)
def toggle_holiday(request):
    try:
        year = int(request.POST.get('year'))
        month = int(request.POST.get('month'))
        day = int(request.POST.get('day'))
        is_holiday = request.POST.get('is_holiday') == 'true'
        
        date_obj = date(year, month, day)
        
        # 更新或创建节假日记录
        holiday, created = Holiday.objects.get_or_create(
            date=date_obj,
            defaults={'description': '手动设置'}
        )
        
        if not is_holiday:
            holiday.delete()
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_POST
@user_passes_test(is_admin)
def regenerate_schedule(request):
    try:
        year = int(request.POST.get('year'))
        month = int(request.POST.get('month'))
        
        # 重新生成排班逻辑
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        
        duty_orders = list(DutyOrder.objects.filter(is_active=True))
        if not duty_orders:
            raise ValueError('没有可用的值班人员')
        
        workday_index = 0
        holiday_index = 0
        
        # 清除已存在的当月排班
        DutySchedule.objects.filter(date__year=year, date__month=month).delete()
        
        current_date = start_date
        while current_date < end_date:
            is_holiday = Holiday.objects.filter(date=current_date).exists()
            is_weekend = current_date.weekday() >= 5
            
            should_schedule = False
            if is_holiday or is_weekend:
                current_index = holiday_index
                should_schedule = True
            elif not is_holiday and current_date.weekday() < 5:
                current_index = workday_index
                should_schedule = True
            
            if should_schedule:
                staff = duty_orders[current_index].staff
                DutySchedule.objects.create(
                    staff=staff,
                    date=current_date,
                    is_holiday=(is_holiday or is_weekend)
                )
                
                if is_holiday or is_weekend:
                    holiday_index = (holiday_index + 1) % len(duty_orders)
                else:
                    workday_index = (workday_index + 1) % len(duty_orders)
            
            current_date += timedelta(days=1)
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@login_required
def duty_list(request):
    schedules = DutySchedule.objects.filter(
        date__month=datetime.now().month
    ).select_related('staff__user').order_by('date')
    
    return render(request, 'duty/duty_list.html', {
        'schedules': schedules
    })

@login_required
def export_schedule(request):
    try:
        # 获取用户指定的年月
        if request.method == 'POST':
            year = int(request.POST.get('year'))
            month = int(request.POST.get('month'))
        else:
            year = datetime.now().year
            month = datetime.now().month

        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}年{month}月值班表"
        
        # 设置列宽
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 20
            
        # 设置标题
        ws.merge_cells('A1:G1')
        title = ws['A1']
        title.value = f"{year}年{month}月值班表"
        title.font = Font(size=16, bold=True)
        title.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置星期标题
        weekdays = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
        header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        weekend_fill = PatternFill(start_color='FFE5E5', end_color='FFE5E5', fill_type='solid')
        holiday_fill = PatternFill(start_color='FFF2F2', end_color='FFF2F2', fill_type='solid')
        workday_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, weekday in enumerate(weekdays, 1):
            cell = ws.cell(row=2, column=col, value=weekday)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if col > 5:  # 周六和周日
                cell.fill = weekend_fill
            else:
                cell.fill = header_fill
                
        # 获取月历数据
        cal = monthcalendar(year, month)
        
        # 获取当月所有排班
        schedules = DutySchedule.objects.filter(
            date__year=year,
            date__month=month
        ).select_related('staff__user')
        
        # 转换为字典格式，包含值班人员和是否节假日的信息
        schedule_dict = {
            schedule.date.day: {
                'staff': schedule.staff.user.username,
                'is_holiday': schedule.is_holiday
            }
            for schedule in schedules
        }
        
        # 获取节假日数据
        holidays = set(Holiday.objects.filter(
            date__year=year,
            date__month=month
        ).values_list('date__day', flat=True))
        
        # 填充日历数据
        current_row = 3
        for week in cal:
            row_height = 60  # 设置行高
            ws.row_dimensions[current_row].height = row_height
            
            for col, day in enumerate(week, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                if day != 0:
                    # 获取当天的排班信息
                    schedule_info = schedule_dict.get(day, {})
                    is_weekend = col > 5
                    is_holiday = day in holidays
                    
                    # 设置日期和值班人员
                    date_text = str(day)
                    duty_staff = schedule_info.get('staff', '')
                    
                    if duty_staff:
                        cell.value = f"{date_text}\n{duty_staff}"
                    else:
                        cell.value = date_text
                    
                    # 根据实际排班状态设置样式
                    if schedule_info.get('is_holiday', False) or is_holiday or is_weekend:
                        cell.font = Font(color='FF0000')  # 红色字体
                        cell.fill = holiday_fill
                    else:
                        cell.font = Font(color='000000')  # 黑色字体
                        cell.fill = workday_fill
            
            current_row += 1
        
        # 设置响应头
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{year}年{month}月值班表.xlsx"'},
        )
        
        # 保存文件
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"导出排班表时发生错误: {str(e)}")
        messages.error(request, '导出排班表时发生错误，请联系管理员')
        return redirect('duty_calendar')

@login_required
@user_passes_test(is_admin)
def manage_duty_order(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        user_id = request.POST.get('user_id')
        
        try:
            if action == 'add':
                user = User.objects.get(id=user_id)
                staff, created = Staff.objects.get_or_create(user=user)
                max_order = DutyOrder.objects.aggregate(Max('order'))['order__max'] or 0
                DutyOrder.objects.create(staff=staff, order=max_order + 1)
                messages.success(request, f'已将 {user.username} 添加到值班顺序')
                
            elif action == 'remove':
                staff = Staff.objects.get(user_id=user_id)
                DutyOrder.objects.filter(staff=staff).update(is_active=False)
                messages.success(request, '已从值班顺序中移除')
                
            elif action == 'reactivate':
                staff = Staff.objects.get(user_id=user_id)
                DutyOrder.objects.filter(staff=staff).update(is_active=True)
                messages.success(request, '已重新激活值班顺序')
                
        except Exception as e:
            messages.error(request, f'操作失败：{str(e)}')
    
    # 获取所有用户和值班顺序
    all_users = User.objects.exclude(is_superuser=True).all()
    duty_orders = DutyOrder.objects.select_related('staff__user').filter(is_active=True)
    inactive_orders = DutyOrder.objects.select_related('staff__user').filter(is_active=False)
    
    # 获取已经在值班顺序中的用户ID列表
    users_in_order = set(DutyOrder.objects.values_list('staff__user_id', flat=True))
    
    # 过滤出未在值班顺序中的用户
    available_users = [user for user in all_users if user.id not in users_in_order]
    
    return render(request, 'duty/manage_duty_order.html', {
        'available_users': available_users,
        'duty_orders': duty_orders,
        'inactive_orders': inactive_orders,
        'current_date': datetime.now(),
    })

@login_required
@user_passes_test(is_admin)
def generate_monthly_schedule(request):
    try:
        year = int(request.POST.get('year'))
        month = int(request.POST.get('month'))
        
        # 重新生成排班逻辑
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        
        duty_orders = list(DutyOrder.objects.filter(is_active=True))
        if not duty_orders:
            raise ValueError('没有可用的值班人员')
        
        workday_index = 0
        holiday_index = 0
        
        # 清除已存在的当月排班
        DutySchedule.objects.filter(date__year=year, date__month=month).delete()
        
        current_date = start_date
        while current_date < end_date:
            is_holiday = Holiday.objects.filter(date=current_date).exists()
            is_weekend = current_date.weekday() >= 5
            
            should_schedule = False
            if is_holiday or is_weekend:
                current_index = holiday_index
                should_schedule = True
            elif not is_holiday and current_date.weekday() < 5:
                current_index = workday_index
                should_schedule = True
            
            if should_schedule:
                staff = duty_orders[current_index].staff
                DutySchedule.objects.create(
                    staff=staff,
                    date=current_date,
                    is_holiday=(is_holiday or is_weekend)
                )
                
                if is_holiday or is_weekend:
                    holiday_index = (holiday_index + 1) % len(duty_orders)
                else:
                    workday_index = (workday_index + 1) % len(duty_orders)
            
            current_date += timedelta(days=1)
        
        messages.success(request, f'已成功生成 {year}年{month}月的值班表')
        return redirect('duty_calendar')
        
    except Exception as e:
        messages.error(request, f'生成值班表时出错：{str(e)}')
        return redirect('duty_calendar')

@login_required
def get_available_swaps(request):
    date_str = request.GET.get('date')
    year, month, day = map(int, date_str.split('-'))
    current_date = date(year, month, day)
    
    # 检查是否是过去的日期
    if current_date < date.today():
        return JsonResponse({'error': '不能交换过去的值班'}, status=400)
    
    # 获取当前用户的值班安排
    current_staff = Staff.objects.get(user=request.user)
    current_duty = DutySchedule.objects.get(
        staff=current_staff,
        date=current_date
    )
    
    # 获取可换班的值班安排（当天及以后的，不包括自己的）
    available_duties = DutySchedule.objects.filter(
        date__gte=date.today(),
        date__year=year,
        date__month=month
    ).exclude(
        staff=current_staff
    ).select_related('staff__user')
    
    duties_data = [{
        'id': duty.id,
        'staff_name': duty.staff.user.username,
        'date': duty.date.strftime('%Y年%m月%d日'),
    } for duty in available_duties]
    
    return JsonResponse(duties_data, safe=False)

@login_required
def request_swap(request):
    if request.method == 'POST':
        try:
            date_str = request.POST.get('requester_duty_date')
            year, month, day = map(int, date_str.split('-'))
            requester_date = date(year, month, day)
            
            # 检查是否是过去的日期
            if requester_date < date.today():
                messages.error(request, '不能交换过去的值班')
                return redirect('duty_calendar')
            
            target_duty_id = request.POST.get('target_duty_id')
            
            requester_staff = Staff.objects.get(user=request.user)
            requester_duty = DutySchedule.objects.get(
                staff=requester_staff,
                date=requester_date
            )
            target_duty = DutySchedule.objects.get(id=target_duty_id)
            
            # 检查目标日期是否是过去的日期
            if target_duty.date < date.today():
                messages.error(request, '不能交换过去的值班')
                return redirect('duty_calendar')
            
            # 检查是否已经存在待处理的换班申请
            existing_request = DutySwapRequest.objects.filter(
                requester_duty=requester_duty,
                status='pending'
            ).exists()
            
            if existing_request:
                messages.error(request, '该值班已有待处理的换班申请')
                return redirect('duty_calendar')
            
            # 创建换班申请
            DutySwapRequest.objects.create(
                requester=requester_staff,
                target=target_duty.staff,
                requester_duty=requester_duty,
                target_duty=target_duty
            )
            
            messages.success(request, '换班申请已发送')
            
        except Exception as e:
            messages.error(request, f'申请换班失败：{str(e)}')
    
    return redirect('duty_calendar')

@login_required
def handle_swap_request(request, request_id):
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            swap_request = DutySwapRequest.objects.get(id=request_id)
            
            if swap_request.target.user != request.user:
                raise PermissionError('您无权处理此换班申请')
            
            if action == 'accept':
                swap_request.accept()
                messages.success(request, '已接受换班申请')
            elif action == 'reject':
                swap_request.reject()
                messages.success(request, '已拒绝换班申请')
            
        except Exception as e:
            messages.error(request, f'处理换班申请失败：{str(e)}')
    
    return redirect('duty_calendar')